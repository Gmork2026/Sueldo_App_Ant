import io
from datetime import date
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.employee import Employee
from app.services.salary import SalaryEngine

engine = SalaryEngine()
VALID_CATEGORIES = engine.CATEGORIES

MONTH_MAP = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}


def normalize_category(raw: str) -> str:
    raw = raw.strip().upper()
    mapping = {
        "VIG. GENERAL": "Vigilador General",
        "VIG.GENERAL": "Vigilador General",
        "VIGILADOR GENERAL": "Vigilador General",
        "VIG. GENERAL + BRIGADISTA": "Vigilador General",
        "VIG. BOMBERO": "Vigilador Bombero",
        "VIGILADOR BOMBERO": "Vigilador Bombero",
        "BOMBERO": "Vigilador Bombero",
        "VIG. ADMINISTRATIVO": "Administrativo",
        "VIG. ADMINISTRATIVA": "Administrativo",
        "ADMINISTRATIVO": "Administrativo",
        "ADMINISTRATIVA": "Administrativo",
        "VIG. PRINCIPAL": "Vigilador Principal",
        "VIG. PRINCIPAL + BRIGADISTA": "Vigilador Principal",
        "VIGILADOR PRINCIPAL": "Vigilador Principal",
        "VIG. MONITOREO": "Operador de Monitoreo",
        "OPERADOR DE MONITOREO": "Operador de Monitoreo",
        "MONITOREO": "Operador de Monitoreo",
        "TECNICO": "Guía Técnico",
        "GUÍA TÉCNICO": "Guía Técnico",
        "GUIA TECNICO": "Guía Técnico",
        "INSTALADOR": "Instalador Sist. Electrónicos",
        "INSTALADOR SIST. ELECTRONICOS": "Instalador Sist. Electrónicos",
        "CADETE": "Vigilador General",
        "0": "",
    }
    return mapping.get(raw, raw.title())


def parse_mes_alta(raw: str | None) -> tuple[int | None, int | None]:
    if not raw:
        return None, None
    text = str(raw).strip().upper()
    parts = text.replace("AÑO", "").strip().split()
    month_num = None
    year_num = None
    for part in parts:
        part_clean = part.strip()
        if part_clean in MONTH_MAP:
            month_num = MONTH_MAP[part_clean]
        elif part_clean.isdigit() and len(part_clean) == 4:
            year_num = int(part_clean)
    return month_num, year_num


def parse_admission_date(row: dict) -> str | None:
    raw = row.get("admission_date", "")
    if raw and hasattr(raw, "strftime"):
        return raw.strftime("%Y-%m-%d")
    if raw and isinstance(raw, str):
        raw_clean = raw.strip()
        if " " in raw_clean:
            raw_clean = raw_clean.split(" ")[0]
        if "-" in raw_clean or "/" in raw_clean:
            return raw_clean

    month, year = parse_mes_alta(row.get("mes_alta"))
    if month and year:
        return f"{year}-{month:02d}-01"
    return None


def parse_row(row_num: int, raw_row: dict) -> dict:
    errors = []
    warnings = []

    name = str(raw_row.get("name", "") or "").strip()
    if not name:
        errors.append("Nombre (VIGILADOR) vacío")

    dni_raw = str(raw_row.get("dni", "") or "").strip()
    dni = ""
    if not dni_raw or dni_raw == "0":
        errors.append("DNI vacío")
    else:
        dni = dni_raw.replace(".", "").replace(",", "").strip()
        if not dni.isdigit() or len(dni) < 7:
            errors.append(f"DNI inválido: '{dni_raw}'")

    legajo_raw = str(raw_row.get("legajo", "") or "").strip()
    legajo = ""
    if legajo_raw:
        legajo = legajo_raw.replace(".", "").replace(",", "").strip()

    category_raw = str(raw_row.get("category", "") or "").strip()
    if not category_raw or category_raw == "0":
        category = "Vigilador General"
        warnings.append("Categoría era 0, se asignó Vigilador General por defecto")
    else:
        category = normalize_category(category_raw)
        if category and category not in VALID_CATEGORIES:
            warnings.append(f"Categoría '{category_raw}' no reconocida, se guardará como '{category}'")

    admission_date = parse_admission_date(raw_row)
    if not admission_date:
        errors.append("Fecha de alta no encontrada (ni F. ALTA ni MES/AÑO ALTA)")

    email = str(raw_row.get("email", "") or "").strip()

    return {
        "row_num": row_num,
        "name": name,
        "legajo": legajo,
        "dni": dni,
        "category": category,
        "admission_date": admission_date,
        "email": email,
        "errors": errors,
        "warnings": warnings,
        "valid": len(errors) == 0,
    }


async def preview_import(file_bytes: bytes, db: AsyncSession) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    COLUMN_MAP = {}
    for i, h in enumerate(headers):
        h_clean = h.strip().upper()
        if h_clean == "LEG":
            COLUMN_MAP[i] = "legajo"
        elif h_clean == "VIGILADOR":
            COLUMN_MAP[i] = "name"
        elif h_clean == "DNI":
            COLUMN_MAP[i] = "dni"
        elif h_clean == "CATEGORIA":
            COLUMN_MAP[i] = "category"
        elif h_clean == "F. ALTA":
            COLUMN_MAP[i] = "admission_date"
        elif h_clean == "MES ALTA":
            COLUMN_MAP[i] = "mes_alta"
        elif h_clean.startswith("A") and "ALTA" in h_clean:
            COLUMN_MAP[i] = "year_alta"
        elif h_clean == "MAIL":
            COLUMN_MAP[i] = "email"
        elif h_clean == "SEXO":
            COLUMN_MAP[i] = "sexo"
        elif h_clean == "ESTADO CIVIL":
            COLUMN_MAP[i] = "estado_civil"
        elif h_clean == "CUIL":
            COLUMN_MAP[i] = "cuil"

    existing_dnis = set()
    result = await db.execute(select(Employee.dni))
    for row in result:
        existing_dnis.add(str(row[0]))

    rows = []
    valid_count = 0
    invalid_count = 0

    for i, excel_row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = {}
        for idx, val in enumerate(excel_row):
            if idx in COLUMN_MAP:
                raw[COLUMN_MAP[idx]] = val

        parsed = parse_row(i, raw)

        if parsed["dni"] in existing_dnis:
            parsed["errors"].append(f"DNI {parsed['dni']} ya existe en la base")
            parsed["valid"] = False
        else:
            existing_dnis.add(parsed["dni"])

        if parsed["valid"]:
            valid_count += 1
        else:
            invalid_count += 1

        rows.append(parsed)

    wb.close()

    return {
        "total_rows": len(rows),
        "valid_count": valid_count,
        "invalid_count": invalid_count,
        "rows": rows,
    }


async def execute_import(file_bytes: bytes, db: AsyncSession, skip_errors: bool = True) -> dict:
    preview = await preview_import(file_bytes, db)

    imported = 0
    skipped = 0
    errors = []

    for row in preview["rows"]:
        if not row["valid"]:
            if skip_errors:
                skipped += 1
                continue
            else:
                errors.append(f"Fila {row['row_num']}: {'; '.join(row['errors'])}")
                continue

        employee = Employee(
            name=row["name"],
            legajo=row["legajo"] or None,
            dni=row["dni"],
            category=row["category"],
            admission_date=row["admission_date"],
        )
        db.add(employee)
        imported += 1

    await db.flush()

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "total": preview["total_rows"],
    }
