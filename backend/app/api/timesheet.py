from datetime import datetime
import calendar
import holidays as ar_holidays
import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.database import get_db
from app.models.timesheet import TimesheetRecord
from app.models.employee import Employee
from app.schemas.timesheet import (
    TimesheetRecordCreate,
    TimesheetRecordUpdate,
    TimesheetRecordRead,
    TimesheetBulkFillRequest,
)
from app.core.deps import CurrentUser, AdminUser

router = APIRouter(prefix="/api/timesheet", tags=["timesheet"])

feriados_arg = ar_holidays.country_holidays("AR")


def calc_hours(entry: str | None, exit: str | None) -> float:
    if not entry or not exit:
        return 0.0
    try:
        fmt = "%H:%M"
        e = datetime.strptime(entry, fmt)
        x = datetime.strptime(exit, fmt)
        diff = (x - e).total_seconds() / 3600
        return round(max(diff, 0), 2)
    except ValueError:
        return 0.0


def detect_holiday(date_str: str) -> tuple[bool, str | None]:
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return False, None
    if d in feriados_arg:
        name = ar_holidays.country_holidays("AR").get(d, None)
        return True, name or "Feriado"
    return False, None


@router.get("/{employee_id}", response_model=list[TimesheetRecordRead])
async def get_timesheet(
    employee_id: int,
    month: int,
    year: int,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    if user.role != "admin":
        from app.models.employee import Employee
        emp_result = await db.execute(
            select(Employee).where(Employee.id == employee_id, Employee.user_id == user.id)
        )
        if not emp_result.scalar_one_or_none():
            raise HTTPException(status_code=403, detail="Acceso denegado")

    month_str = f"{year}-{month:02d}"
    result = await db.execute(
        select(TimesheetRecord)
        .where(
            TimesheetRecord.employee_id == employee_id,
            TimesheetRecord.date.like(f"{month_str}%"),
        )
        .order_by(TimesheetRecord.date)
    )
    return result.scalars().all()


@router.post("", response_model=TimesheetRecordRead)
async def create_or_update_record(
    data: TimesheetRecordCreate,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    if user.role != "admin":
        from app.models.employee import Employee
        emp_result = await db.execute(
            select(Employee).where(
                Employee.id == data.employee_id, Employee.user_id == user.id
            )
        )
        if not emp_result.scalar_one_or_none():
            raise HTTPException(status_code=403, detail="Acceso denegado")

    result = await db.execute(
        select(TimesheetRecord).where(
            TimesheetRecord.employee_id == data.employee_id,
            TimesheetRecord.date == data.date,
        )
    )
    existing = result.scalar_one_or_none()

    update_data = data.model_dump()

    if not data.is_franco and data.entry_time and data.exit_time and not data.total_hours:
        update_data["total_hours"] = calc_hours(data.entry_time, data.exit_time)

    if not data.is_holiday and not data.is_franco:
        is_h, h_name = detect_holiday(data.date)
        if is_h:
            update_data["is_holiday"] = True
            update_data["holiday_name"] = h_name

    if existing:
        for field, value in update_data.items():
            if field != "employee_id":
                setattr(existing, field, value)
        await db.flush()
        await db.refresh(existing)
        return existing

    record = TimesheetRecord(**update_data)
    db.add(record)
    await db.flush()
    await db.refresh(record)
    return record


@router.put("/{record_id}", response_model=TimesheetRecordRead)
async def update_record(
    record_id: int,
    data: TimesheetRecordUpdate,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(TimesheetRecord).where(TimesheetRecord.id == record_id)
    )
    record = result.scalar_one_or_none()

    if not record:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    if user.role != "admin":
        from app.models.employee import Employee
        emp_result = await db.execute(
            select(Employee).where(
                Employee.id == record.employee_id, Employee.user_id == user.id
            )
        )
        if not emp_result.scalar_one_or_none():
            raise HTTPException(status_code=403, detail="Acceso denegado")

    update_data = data.model_dump(exclude_unset=True)

    entry = update_data.get("entry_time", record.entry_time)
    exit_t = update_data.get("exit_time", record.exit_time)
    is_franco = update_data.get("is_franco", record.is_franco)
    if not is_franco and entry and exit_t and "total_hours" not in update_data:
        update_data["total_hours"] = calc_hours(entry, exit_t)

    for field, value in update_data.items():
        setattr(record, field, value)

    await db.flush()
    await db.refresh(record)
    return record


@router.delete("/{record_id}")
async def delete_record(record_id: int, user: CurrentUser, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(TimesheetRecord).where(TimesheetRecord.id == record_id)
    )
    record = result.scalar_one_or_none()

    if not record:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    if user.role != "admin":
        from app.models.employee import Employee
        emp_result = await db.execute(
            select(Employee).where(
                Employee.id == record.employee_id, Employee.user_id == user.id
            )
        )
        if not emp_result.scalar_one_or_none():
            raise HTTPException(status_code=403, detail="Acceso denegado")

    await db.delete(record)
    await db.flush()
    return {"ok": True}


@router.post("/bulk", response_model=list[TimesheetRecordRead])
async def bulk_fill_month(
    data: TimesheetBulkFillRequest,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    if user.role != "admin":
        from app.models.employee import Employee
        emp_result = await db.execute(
            select(Employee).where(
                Employee.id == data.employee_id, Employee.user_id == user.id
            )
        )
        if not emp_result.scalar_one_or_none():
            raise HTTPException(status_code=403, detail="Acceso denegado")

    num_days = calendar.monthrange(data.year, data.month)[1]

    month_str = f"{data.year}-{data.month:02d}"
    existing_result = await db.execute(
        select(TimesheetRecord.date).where(
            TimesheetRecord.employee_id == data.employee_id,
            TimesheetRecord.date.like(f"{month_str}%"),
        )
    )
    existing_dates = {row[0] for row in existing_result.all()}

    skip_dates = set(data.skip_dates or [])
    created = []

    for day in range(1, num_days + 1):
        date_str = f"{data.year}-{data.month:02d}-{day:02d}"

        if date_str in existing_dates or date_str in skip_dates:
            continue

        d = datetime.strptime(date_str, "%Y-%m-%d").date()
        if d.weekday() == 6:
            continue

        is_h, h_name = detect_holiday(date_str)

        total = calc_hours(data.entry_time, data.exit_time)

        record = TimesheetRecord(
            employee_id=data.employee_id,
            date=date_str,
            entry_time=data.entry_time,
            exit_time=data.exit_time,
            total_hours=total,
            is_franco=False,
            is_holiday=is_h,
            holiday_name=h_name,
        )
        db.add(record)
        created.append(record)

    await db.flush()
    for r in created:
        await db.refresh(r)
    return created


MONTH_NAMES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]
DAY_NAMES = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]


@router.get("/{employee_id}/export")
async def export_timesheet_excel(
    employee_id: int,
    month: int,
    year: int,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    if user.role != "admin":
        emp_result = await db.execute(
            select(Employee).where(Employee.id == employee_id, Employee.user_id == user.id)
        )
        if not emp_result.scalar_one_or_none():
            raise HTTPException(status_code=403, detail="Acceso denegado")

    emp_result = await db.execute(select(Employee).where(Employee.id == employee_id))
    employee = emp_result.scalar_one_or_none()
    if not employee:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")

    month_str = f"{year}-{month:02d}"
    result = await db.execute(
        select(TimesheetRecord)
        .where(
            TimesheetRecord.employee_id == employee_id,
            TimesheetRecord.date.like(f"{month_str}%"),
        )
        .order_by(TimesheetRecord.date)
    )
    records = {r.date: r for r in result.scalars().all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Horas"

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 30

    ws.merge_cells("A1:I1")
    ws["A1"] = f"Control de Horas - {MONTH_NAMES[month - 1]} {year}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Empleado:"
    ws["A3"].font = header_font
    ws["B3"] = employee.name
    ws["A4"] = "DNI:"
    ws["A4"].font = header_font
    ws["B4"] = employee.dni
    ws["A5"] = "Categoría:"
    ws["A5"].font = header_font
    ws["B5"] = employee.category
    if employee.legajo:
        ws["D3"] = "Legajo:"
        ws["D3"].font = header_font
        ws["E3"] = employee.legajo

    headers = ["Día", "Fecha", "Entrada", "Salida", "Horas", "Tipo", "Diurnas", "Nocturnas", "Notas"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    num_days = calendar.monthrange(year, month)[1]
    row = 8
    total_hours = 0.0
    total_diurnas = 0.0
    total_nocturnas = 0.0
    total_work_days = 0

    for day in range(1, num_days + 1):
        ds = f"{year}-{month:02d}-{day:02d}"
        d = datetime.strptime(ds, "%Y-%m-%d").date()
        rec = records.get(ds)

        ws.cell(row=row, column=1, value=DAY_NAMES[d.weekday()]).border = thin_border
        ws.cell(row=row, column=2, value=d.strftime("%d/%m/%Y")).border = thin_border

        if rec:
            ws.cell(row=row, column=3, value=rec.entry_time or "").border = thin_border
            ws.cell(row=row, column=4, value=rec.exit_time or "").border = thin_border

            hrs = float(rec.total_hours)
            total_hours += hrs
            ws.cell(row=row, column=5, value=hrs).border = thin_border
            ws.cell(row=row, column=5).number_format = "0.00"

            if rec.is_franco:
                ws.cell(row=row, column=6, value="Franco").border = thin_border
                ws.cell(row=row, column=6).font = Font(color="7030A0", bold=True)
                ws.cell(row=row, column=7, value=0).border = thin_border
                ws.cell(row=row, column=8, value=0).border = thin_border
            elif rec.is_holiday:
                ws.cell(row=row, column=6, value=rec.holiday_name or "Feriado").border = thin_border
                ws.cell(row=row, column=6).font = Font(color="C55A11", bold=True)
                if hrs > 0:
                    total_work_days += 1
                if hrs > 0 and rec.entry_time and rec.exit_time:
                    diurnas, nocturnas = _calc_diurnas_nocturnas(rec.entry_time, rec.exit_time)
                    total_diurnas += diurnas
                    total_nocturnas += nocturnas
                    ws.cell(row=row, column=7, value=round(diurnas, 2)).border = thin_border
                    ws.cell(row=row, column=7).number_format = "0.00"
                    ws.cell(row=row, column=8, value=round(nocturnas, 2)).border = thin_border
                    ws.cell(row=row, column=8).number_format = "0.00"
                else:
                    ws.cell(row=row, column=7, value=0).border = thin_border
                    ws.cell(row=row, column=8, value=0).border = thin_border
            elif hrs > 0:
                ws.cell(row=row, column=6, value="Trabajó").border = thin_border
                ws.cell(row=row, column=6).font = Font(color="006100")
                total_work_days += 1

                if rec.entry_time and rec.exit_time:
                    diurnas, nocturnas = _calc_diurnas_nocturnas(rec.entry_time, rec.exit_time)
                    total_diurnas += diurnas
                    total_nocturnas += nocturnas
                    ws.cell(row=row, column=7, value=round(diurnas, 2)).border = thin_border
                    ws.cell(row=row, column=7).number_format = "0.00"
                    ws.cell(row=row, column=8, value=round(nocturnas, 2)).border = thin_border
                    ws.cell(row=row, column=8).number_format = "0.00"
                else:
                    ws.cell(row=row, column=7, value=0).border = thin_border
                    ws.cell(row=row, column=8, value=0).border = thin_border
            else:
                ws.cell(row=row, column=6, value="").border = thin_border
                ws.cell(row=row, column=7, value=0).border = thin_border
                ws.cell(row=row, column=8, value=0).border = thin_border

            ws.cell(row=row, column=9, value=rec.notes or "").border = thin_border
        else:
            for c in range(3, 10):
                ws.cell(row=row, column=c, value="" if c != 7 and c != 8 else 0).border = thin_border
            ws.cell(row=row, column=5, value=0).border = thin_border
            ws.cell(row=row, column=5).number_format = "0.00"

        row += 1

    total_row_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    ws.cell(row=row, column=1, value="TOTALES").font = total_row_font
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).fill = total_fill
    ws.cell(row=row, column=3).border = thin_border
    ws.cell(row=row, column=3).fill = total_fill
    ws.cell(row=row, column=4).border = thin_border
    ws.cell(row=row, column=4).fill = total_fill

    c = ws.cell(row=row, column=5, value=round(total_hours, 2))
    c.font = total_row_font
    c.fill = total_fill
    c.border = thin_border
    c.number_format = "0.00"

    ws.cell(row=row, column=6, value=f"{total_work_days} días").font = total_row_font
    ws.cell(row=row, column=6).fill = total_fill
    ws.cell(row=row, column=6).border = thin_border

    c = ws.cell(row=row, column=7, value=round(total_diurnas, 2))
    c.font = total_row_font
    c.fill = total_fill
    c.border = thin_border
    c.number_format = "0.00"

    c = ws.cell(row=row, column=8, value=round(total_nocturnas, 2))
    c.font = total_row_font
    c.fill = total_fill
    c.border = thin_border
    c.number_format = "0.00"

    ws.cell(row=row, column=9).border = thin_border
    ws.cell(row=row, column=9).fill = total_fill

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"fichadas_{employee.name.replace(' ', '_')}_{month:02d}_{year}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


DIURNA_START = 360  # 06:00
DIURNA_END = 1260   # 21:00


def _calc_diurnas_nocturnas(entry: str, exit_time: str) -> tuple[float, float]:
    def to_min(t: str) -> int:
        h, m = t.split(":")
        return int(h) * 60 + int(m)

    start = to_min(entry)
    end = to_min(exit_time)
    if end <= start:
        end += 24 * 60

    diurnas_min = 0
    nocturnas_min = 0
    cursor = start

    while cursor < end:
        pos = cursor % (24 * 60)
        if pos < DIURNA_START:
            boundary = min(end, cursor + (DIURNA_START - pos))
        elif pos < DIURNA_END:
            boundary = min(end, cursor + (DIURNA_END - pos))
        else:
            boundary = min(end, cursor + (24 * 60 - pos))

        if DIURNA_START <= pos < DIURNA_END:
            diurnas_min += boundary - cursor
        else:
            nocturnas_min += boundary - cursor
        cursor = boundary

    return diurnas_min / 60, nocturnas_min / 60
