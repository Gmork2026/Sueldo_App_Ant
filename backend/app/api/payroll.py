from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import io
from openpyxl import Workbook

from app.database import get_db
from app.models.employee import Employee
from app.models.timesheet import TimesheetRecord
from app.models.payroll import Payroll
from app.schemas.payroll import PayrollCalculateRequest, PayrollRead
from app.core.deps import CurrentUser, AdminUser

router = APIRouter(prefix="/api/payroll", tags=["payroll"])


@router.post("/calculate", response_model=PayrollRead)
async def calculate_payroll(
    data: PayrollCalculateRequest,
    admin: AdminUser,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Employee).where(Employee.id == data.employee_id))
    employee = result.scalar_one_or_none()
    if not employee:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")

    from app.services.salary import SalaryEngine

    engine = SalaryEngine()

    month_str = f"{data.year}-{data.month:02d}"
    ts_result = await db.execute(
        select(TimesheetRecord).where(
            TimesheetRecord.employee_id == data.employee_id,
            TimesheetRecord.date.like(f"{month_str}%"),
        )
    )
    records = ts_result.scalars().all()

    total_hours = sum(float(r.total_hours) for r in records)

    payroll_data = engine.calculate(
        category=employee.category,
        year=data.year,
        month=data.month,
        admission_date=employee.admission_date,
        total_hours=total_hours,
        dias_vacaciones=data.dias_vacaciones,
    )

    existing_result = await db.execute(
        select(Payroll).where(
            Payroll.employee_id == data.employee_id,
            Payroll.month == data.month,
            Payroll.year == data.year,
        )
    )
    existing = existing_result.scalar_one_or_none()

    if existing:
        for field, value in payroll_data.items():
            setattr(existing, field, value)
        await db.flush()
        await db.refresh(existing)
        return existing

    payroll = Payroll(employee_id=data.employee_id, **payroll_data)
    db.add(payroll)
    await db.flush()
    await db.refresh(payroll)
    return payroll


@router.get("/{employee_id}", response_model=list[PayrollRead])
async def get_payroll(
    employee_id: int,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    if user.role != "admin":
        emp_result = await db.execute(
            select(Employee).where(
                Employee.id == employee_id, Employee.user_id == user.id
            )
        )
        if not emp_result.scalar_one_or_none():
            raise HTTPException(status_code=403, detail="Acceso denegado")

    result = await db.execute(
        select(Payroll).where(Payroll.employee_id == employee_id)
    )
    return result.scalars().all()


@router.get("/{payroll_id}/export")
async def export_payroll_excel(
    payroll_id: int,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Payroll).where(Payroll.id == payroll_id))
    payroll = result.scalar_one_or_none()
    if not payroll:
        raise HTTPException(status_code=404, detail="Liquidación no encontrada")

    if user.role != "admin":
        emp_result = await db.execute(
            select(Employee).where(
                Employee.id == payroll.employee_id, Employee.user_id == user.id
            )
        )
        if not emp_result.scalar_one_or_none():
            raise HTTPException(status_code=403, detail="Acceso denegado")

    emp_result = await db.execute(
        select(Employee).where(Employee.id == payroll.employee_id)
    )
    employee = emp_result.scalar_one()

    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidación"

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    money_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    green_font = Font(bold=True, color="006100", size=11)
    red_font = Font(bold=True, color="9C0006", size=11)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18

    ws.merge_cells("A1:C1")
    ws["A1"] = f"Liquidación de Sueldo"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Empleado:"
    ws["A3"].font = header_font
    ws["B3"] = employee.name
    ws["A4"] = "Categoría:"
    ws["A4"].font = header_font
    ws["B4"] = employee.category
    ws["A5"] = "Período:"
    ws["A5"].font = header_font
    ws["B5"] = f"{MONTH_NAMES[payroll.month - 1]} {payroll.year}"

    MONTH_NAMES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                   "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    row = 7
    for col_letter in ["A", "B", "C"]:
        cell = ws.cell(row=row, column=ord(col_letter) - ord("A") + 1)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border
    ws.cell(row=row, column=1, value="Concepto")
    ws.cell(row=row, column=2, value="Detalle")
    ws.cell(row=row, column=3, value="Importe")

    def add_row(label, detail, amount):
        nonlocal row
        row += 1
        ws.cell(row=row, column=1, value=label).border = thin_border
        ws.cell(row=row, column=2, value=detail).border = thin_border
        c = ws.cell(row=row, column=3, value=amount)
        c.number_format = money_fmt
        c.border = thin_border
        c.alignment = Alignment(horizontal="right")

    add_row("Básico", "", float(payroll.basic_salary))
    add_row("Antigüedad", f"{payroll.seniority_years} años", float(payroll.seniority_amount))
    add_row("Presentismo", "", float(payroll.presentismo))
    add_row("Horas Extras", f"{payroll.overtime_hours} hs", float(payroll.overtime_amount))
    add_row("Feriados", f"{payroll.holiday_hours} días", float(payroll.holiday_amount))
    add_row("Nocturnidad", f"{payroll.night_hours} hs", float(payroll.night_amount))

    row += 1
    ws.cell(row=row, column=1, value="").border = thin_border
    ws.cell(row=row, column=2, value="").border = thin_border
    ws.cell(row=row, column=3, value="").border = thin_border

    add_row("Viáticos", "", float(payroll.viaticos))
    add_row("No Remunerativo", "", float(payroll.non_remunerative))

    row += 1
    for c in range(1, 4):
        ws.cell(row=row, column=c).border = thin_border

    row += 1
    ws.cell(row=row, column=1, value="BRUTO").font = Font(bold=True, size=11)
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2).border = thin_border
    c = ws.cell(row=row, column=3, value=float(payroll.gross_salary))
    c.font = Font(bold=True, size=11)
    c.number_format = money_fmt
    c.border = thin_border
    c.alignment = Alignment(horizontal="right")

    row += 1
    ws.cell(row=row, column=1, value="Deducciones").font = red_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2).border = thin_border
    c = ws.cell(row=row, column=3, value=float(payroll.deductions))
    c.font = red_font
    c.number_format = money_fmt
    c.border = thin_border
    c.alignment = Alignment(horizontal="right")

    row += 1
    ws.cell(row=row, column=1, value="NETO A COBRAR").font = green_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2).border = thin_border
    c = ws.cell(row=row, column=3, value=float(payroll.net_salary))
    c.font = green_font
    c.number_format = money_fmt
    c.border = thin_border
    c.alignment = Alignment(horizontal="right")

    if payroll.sac_bruto > 0:
        row += 2
        ws.cell(row=row, column=1, value="SAC (Aguinaldo)").font = Font(bold=True, size=12)
        row += 1
        add_row("SAC Bruto", "", float(payroll.sac_bruto))
        add_row("SAC Deducciones", "", float(payroll.sac_deducciones))
        add_row("SAC Neto", "", float(payroll.sac_neto))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"liquidacion_{employee.name.replace(' ', '_')}_{payroll.month:02d}_{payroll.year}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
